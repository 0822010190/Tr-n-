"""
Trộn Đề Word Online - AIOMT Premium (Customized) - XLSX OUTPUT
Streamlit App

TÍNH NĂNG:
1) {{MA_DE}} trong TextBox (header/footer/document) -> tự thay theo lựa chọn trên web:
   - Đầy đủ 3 số (101)
   - 2 số đầu (10)
   - 2 số cuối (01)
   Ngoài ra hỗ trợ token cố định: {{MA_DE_2DAU}}, {{MA_DE_2CUOI}}
2) PHẦN 1 & 2: đáp án đúng xác định bằng gạch chân (underline) nhưng ĐỀ TRỘN XONG sẽ bỏ gạch chân
3) PHẦN 3: lấy đáp án theo dòng "Đáp án: ..." nhưng ĐỀ TRỘN XONG sẽ xóa dòng đó
4) Xuất 1 file đáp án duy nhất dạng bảng EXCEL .XLSX:
   - merge nhóm tiêu đề
   - kẻ bảng, căn giữa, freeze panes
5) PHẦN 2 đúng/sai xuất dạng "ĐSĐS" (không dấu phẩy)
"""

import streamlit as st
import re
import random
import zipfile
import io
from xml.dom import minidom

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ==================== PAGE CONFIG ====================

st.set_page_config(
    page_title="Trộn đề Word - AIOMT (XLSX)",
    page_icon="🎲",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==================== UI STYLE ====================

st.markdown(
    """
<style>
:root{
  --primary:#0d9488;
  --primary2:#14b8a6;
  --bg:#f6fbfb;
  --card:#ffffff;
  --text:#0f172a;
  --muted:#64748b;
}

.stApp { background: var(--bg); }
.block-container { padding-top: 1.2rem; padding-bottom: 2.5rem; }

.hero{
  background: linear-gradient(90deg, rgba(13,148,136,0.12), rgba(20,184,166,0.12));
  border: 1px solid rgba(13,148,136,0.25);
  border-radius: 18px;
  padding: 18px 18px;
  margin-bottom: 14px;
}
.hero h1{ margin:0; color: var(--text); font-size: 28px; }
.hero p{ margin:6px 0 0 0; color: var(--muted); }

.card{
  background: var(--card);
  border: 1px solid rgba(2,132,199,0.15);
  border-radius: 16px;
  padding: 16px 16px;
  box-shadow: 0 6px 20px rgba(2,132,199,0.06);
}

.badge{
  display:inline-block;
  padding: 3px 10px;
  border-radius: 999px;
  background: rgba(13,148,136,0.12);
  border: 1px solid rgba(13,148,136,0.25);
  color: var(--primary);
  font-size: 12px;
  margin-right: 6px;
}

.small{
  color: var(--muted);
  font-size: 13px;
}

.hr{
  height:1px;
  background: rgba(100,116,139,0.18);
  margin: 12px 0;
}

.stButton > button{
  width:100%;
  background: linear-gradient(90deg, var(--primary), var(--primary2));
  color:#fff;
  border:0;
  border-radius: 12px;
  padding: 0.85rem 1rem;
  font-weight: 700;
  font-size: 1.05rem;
}
.stButton > button:hover{
  filter: brightness(0.95);
  box-shadow: 0 10px 24px rgba(13,148,136,0.22);
}

footer{
  color: var(--muted);
  font-size: 12.5px;
  text-align:center;
  margin-top: 18px;
}
</style>
""",
    unsafe_allow_html=True
)

# ==================== WORD XML HELPERS ====================

W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"


def shuffle_array(arr):
    out = arr.copy()
    for i in range(len(out) - 1, 0, -1):
        j = random.randint(0, i)
        out[i], out[j] = out[j], out[i]
    return out


def get_text(block):
    texts = []
    t_nodes = block.getElementsByTagNameNS(W_NS, "t")
    for t in t_nodes:
        if t.firstChild and t.firstChild.nodeValue:
            texts.append(t.firstChild.nodeValue)
    return "".join(texts).strip()


def _run_text(run):
    texts = []
    t_nodes = run.getElementsByTagNameNS(W_NS, "t")
    for t in t_nodes:
        if t.firstChild and t.firstChild.nodeValue:
            texts.append(t.firstChild.nodeValue)
    return "".join(texts)


def _is_label_only_text(s: str) -> bool:
    t = (s or "").strip()
    if re.fullmatch(r"[A-D]\.", t):
        return True
    if re.fullmatch(r"[A-D]\)", t):
        return True
    if re.fullmatch(r"[a-d]\)", t):
        return True
    if re.fullmatch(r"Câu\s*\d+\.", t, flags=re.IGNORECASE):
        return True
    return False


def run_has_underline(run) -> bool:
    rPr_list = run.getElementsByTagNameNS(W_NS, "rPr")
    if not rPr_list:
        return False
    rPr = rPr_list[0]
    u_list = rPr.getElementsByTagNameNS(W_NS, "u")
    if not u_list:
        return False
    u_el = u_list[0]
    val = u_el.getAttributeNS(W_NS, "val") or u_el.getAttribute("w:val") or ""
    val = (val or "").strip().lower()
    if val == "none":
        return False
    return True


def block_has_underlined_content(block) -> bool:
    r_nodes = block.getElementsByTagNameNS(W_NS, "r")
    for r in r_nodes:
        if not run_has_underline(r):
            continue
        rt = _run_text(r)
        if _is_label_only_text(rt):
            continue
        if (rt or "").strip():
            return True
    return False


def remove_underline_in_block(block):
    """Bỏ gạch chân trong 1 block để đề trộn không lộ đáp án."""
    r_nodes = block.getElementsByTagNameNS(W_NS, "r")
    for r in r_nodes:
        rPr_list = r.getElementsByTagNameNS(W_NS, "rPr")
        if not rPr_list:
            continue
        rPr = rPr_list[0]
        u_list = rPr.getElementsByTagNameNS(W_NS, "u")
        for u in u_list:
            u.setAttribute("w:val", "none")


def remove_short_answer_lines(question_blocks):
    """Xóa các block dạng 'Đáp án: ...' khỏi PHẦN 3 trong đề trộn."""
    out = []
    for b in question_blocks:
        txt = get_text(b)
        if re.match(r"^\s*Đáp\s*án\s*[:\-]\s*.+\s*$", txt, flags=re.IGNORECASE):
            continue
        out.append(b)
    return out


def style_run_blue_bold(run):
    """Tô xanh + đậm cho nhãn (chỉ để đẹp đề)."""
    doc = run.ownerDocument

    rPr_list = run.getElementsByTagNameNS(W_NS, "rPr")
    if rPr_list:
        rPr = rPr_list[0]
    else:
        rPr = doc.createElementNS(W_NS, "w:rPr")
        run.insertBefore(rPr, run.firstChild)

    color_list = rPr.getElementsByTagNameNS(W_NS, "color")
    if color_list:
        color_el = color_list[0]
    else:
        color_el = doc.createElementNS(W_NS, "w:color")
        rPr.appendChild(color_el)
    color_el.setAttributeNS(W_NS, "w:val", "0000FF")

    b_list = rPr.getElementsByTagNameNS(W_NS, "b")
    if not b_list:
        b_el = doc.createElementNS(W_NS, "w:b")
        rPr.appendChild(b_el)


def update_mcq_label(paragraph, new_label):
    t_nodes = paragraph.getElementsByTagNameNS(W_NS, "t")
    if not t_nodes:
        return

    new_letter = new_label[0].upper()
    new_punct = "."

    for i, t in enumerate(t_nodes):
        if not t.firstChild or not t.firstChild.nodeValue:
            continue

        txt = t.firstChild.nodeValue
        m = re.match(r"^(\s*)([A-D])([\.\)])?", txt, re.IGNORECASE)
        if not m:
            continue

        leading_space = m.group(1) or ""
        old_punct = m.group(3) or ""
        after_match = txt[m.end():]

        if old_punct:
            t.firstChild.nodeValue = leading_space + new_letter + new_punct + after_match
        else:
            t.firstChild.nodeValue = leading_space + new_letter + after_match
            found_punct = False
            for j in range(i + 1, len(t_nodes)):
                t2 = t_nodes[j]
                if not t2.firstChild or not t2.firstChild.nodeValue:
                    continue
                txt2 = t2.firstChild.nodeValue
                if re.match(r"^[\.\)]", txt2):
                    t2.firstChild.nodeValue = new_punct + txt2[1:]
                    found_punct = True
                    break
                elif re.match(r"^\s*$", txt2):
                    continue
                else:
                    break
            if not found_punct:
                t.firstChild.nodeValue = leading_space + new_letter + new_punct + after_match

        run = t.parentNode
        if run and run.localName == "r":
            style_run_blue_bold(run)
        break


def update_tf_label(paragraph, new_label):
    t_nodes = paragraph.getElementsByTagNameNS(W_NS, "t")
    if not t_nodes:
        return

    new_letter = new_label[0].lower()
    new_punct = ")"

    for i, t in enumerate(t_nodes):
        if not t.firstChild or not t.firstChild.nodeValue:
            continue

        txt = t.firstChild.nodeValue
        m = re.match(r"^(\s*)([a-d])(\))?", txt, re.IGNORECASE)
        if not m:
            continue

        leading_space = m.group(1) or ""
        old_punct = m.group(3) or ""
        after_match = txt[m.end():]

        if old_punct:
            t.firstChild.nodeValue = leading_space + new_letter + new_punct + after_match
        else:
            t.firstChild.nodeValue = leading_space + new_letter + after_match
            found_punct = False
            for j in range(i + 1, len(t_nodes)):
                t2 = t_nodes[j]
                if not t2.firstChild or not t2.firstChild.nodeValue:
                    continue
                txt2 = t2.firstChild.nodeValue
                if re.match(r"^\)", txt2):
                    found_punct = True
                    break
                elif re.match(r"^\s*$", txt2):
                    continue
                else:
                    break
            if not found_punct:
                t.firstChild.nodeValue = leading_space + new_letter + new_punct + after_match

        run = t.parentNode
        if run and run.localName == "r":
            style_run_blue_bold(run)
        break


def update_question_label(paragraph, new_label):
    t_nodes = paragraph.getElementsByTagNameNS(W_NS, "t")
    if not t_nodes:
        return

    for i, t in enumerate(t_nodes):
        if not t.firstChild or not t.firstChild.nodeValue:
            continue

        txt = t.firstChild.nodeValue
        m = re.match(r"^(\s*)(Câu\s*)(\d+)(\.)?", txt, re.IGNORECASE)
        if not m:
            continue

        leading_space = m.group(1) or ""
        after_match = txt[m.end():]

        t.firstChild.nodeValue = leading_space + new_label + after_match

        run = t.parentNode
        if run and run.localName == "r":
            style_run_blue_bold(run)

        for j in range(i + 1, len(t_nodes)):
            t2 = t_nodes[j]
            if not t2.firstChild or not t2.firstChild.nodeValue:
                continue
            txt2 = t2.firstChild.nodeValue
            if re.match(r"^[\s0-9\.]*$", txt2) and txt2.strip():
                t2.firstChild.nodeValue = ""
            elif re.match(r"^\s*$", txt2):
                continue
            else:
                break
        break


def find_part_index(blocks, part_number):
    pattern = re.compile(rf"PHẦN\s*{part_number}\b", re.IGNORECASE)
    for i, block in enumerate(blocks):
        text = get_text(block)
        if pattern.search(text):
            return i
    return -1


def parse_questions_in_range(blocks, start, end):
    part_blocks = blocks[start:end]
    intro = []
    questions = []

    i = 0
    while i < len(part_blocks):
        text = get_text(part_blocks[i])
        if re.match(r"^Câu\s*\d+\b", text):
            break
        intro.append(part_blocks[i])
        i += 1

    while i < len(part_blocks):
        text = get_text(part_blocks[i])
        if re.match(r"^Câu\s*\d+\b", text):
            group = [part_blocks[i]]
            i += 1
            while i < len(part_blocks):
                t2 = get_text(part_blocks[i])
                if re.match(r"^Câu\s*\d+\b", t2):
                    break
                if re.match(r"^PHẦN\s*\d\b", t2, re.IGNORECASE):
                    break
                group.append(part_blocks[i])
                i += 1
            questions.append(group)
        else:
            intro.append(part_blocks[i])
            i += 1

    return intro, questions


# ==================== PLACEHOLDER MA_DE (WITH MODE) ====================

def replace_ma_de_placeholders(xml_text: str, ma_de: int, ma_de_mode: str = "full") -> str:
    """
    Token hỗ trợ:
    - {{MA_DE}}: theo mode (full | 2dau | 2cuoi)
    - {{MA_DE_2DAU}}: luôn 2 số đầu
    - {{MA_DE_2CUOI}}: luôn 2 số cuối (giữ 0 nếu có)
    """
    s = xml_text

    code3 = f"{int(ma_de):03d}"  # luôn 3 chữ số
    first2 = code3[:2]
    last2 = code3[-2:]

    if ma_de_mode == "2dau":
        main_value = first2
    elif ma_de_mode == "2cuoi":
        main_value = last2
    else:
        main_value = code3

    s = s.replace("{{MA_DE}}", main_value).replace("{MA_DE}", main_value)
    s = s.replace("{{MA_DE_2DAU}}", first2).replace("{MA_DE_2DAU}", first2)
    s = s.replace("{{MA_DE_2CUOI}}", last2).replace("{MA_DE_2CUOI}", last2)
    return s


# ==================== PART 1: MCQ ====================

def shuffle_mcq_options(question_blocks):
    indices = []
    for i, block in enumerate(question_blocks):
        text = get_text(block)
        if re.match(r"^\s*[A-D][\.\)]", text, re.IGNORECASE):
            indices.append(i)

    if len(indices) < 2:
        return question_blocks, ""

    options = []
    correct_old = ""
    for idx in indices:
        block = question_blocks[idx]
        txt = get_text(block).strip()
        old_letter = txt[0].upper() if txt else ""
        is_correct = block_has_underlined_content(block)
        if is_correct:
            correct_old = old_letter
        options.append((old_letter, block, is_correct))

    shuffled = shuffle_array(options)

    new_correct = ""
    for new_pos, (old_letter, _, _) in enumerate(shuffled):
        if old_letter == correct_old and correct_old:
            new_correct = chr(ord("A") + new_pos)
            break

    min_idx = min(indices)
    max_idx = max(indices)
    before = question_blocks[:min_idx]
    after = question_blocks[max_idx + 1:]

    new_blocks = before + [b for _, b, _ in shuffled] + after

    # bỏ gạch chân sau khi đã lấy đáp án
    for _, b, _ in shuffled:
        remove_underline_in_block(b)

    return new_blocks, new_correct


def relabel_mcq_options(question_blocks):
    letters = ["A", "B", "C", "D"]
    option_blocks = []
    for block in question_blocks:
        text = get_text(block)
        if re.match(r"^\s*[A-D][\.\)]", text, re.IGNORECASE):
            option_blocks.append(block)

    for idx, block in enumerate(option_blocks):
        letter = letters[idx] if idx < len(letters) else letters[-1]
        update_mcq_label(block, f"{letter}.")


# ==================== PART 2: TRUE/FALSE ====================

def shuffle_tf_options_and_key(question_blocks):
    option_map = {}
    for i, block in enumerate(question_blocks):
        text = get_text(block)
        m = re.match(r"^\s*([a-d])\)", text, re.IGNORECASE)
        if m:
            letter = m.group(1).lower()
            truth = block_has_underlined_content(block)
            option_map[letter] = (i, block, truth)

    if len(option_map) < 2:
        key_labels = []
        for k in ["a", "b", "c", "d"]:
            if k in option_map:
                key_labels.append("Đ" if option_map[k][2] else "S")
                remove_underline_in_block(option_map[k][1])
            else:
                key_labels.append("")
        return question_blocks, "".join(key_labels).strip()

    abc = []
    for k in ["a", "b", "c"]:
        if k in option_map:
            abc.append((option_map[k][1], option_map[k][2]))

    abc_shuffled = shuffle_array(abc) if len(abc) >= 2 else abc

    d_block = option_map["d"][1] if "d" in option_map else None
    d_truth = option_map["d"][2] if "d" in option_map else None

    all_idx = [v[0] for v in option_map.values()]
    min_idx = min(all_idx)
    max_idx = max(all_idx)

    before = question_blocks[:min_idx]
    after = question_blocks[max_idx + 1:]

    middle_blocks = [b for (b, _) in abc_shuffled]
    middle_truths = [t for (_, t) in abc_shuffled]

    if d_block is not None:
        middle_blocks.append(d_block)
        middle_truths.append(d_truth)

    new_blocks = before + middle_blocks + after

    key_labels = [("Đ" if t else "S") for t in middle_truths]
    while len(key_labels) < 4:
        key_labels.append("")
    key_str = "".join(key_labels[:4])  # ĐSĐS

    # bỏ gạch chân sau khi đã lấy key
    for b in middle_blocks:
        remove_underline_in_block(b)

    return new_blocks, key_str


def relabel_tf_options(question_blocks):
    letters = ["a", "b", "c", "d"]
    option_blocks = []
    for block in question_blocks:
        text = get_text(block)
        if re.match(r"^\s*[a-d]\)", text, re.IGNORECASE):
            option_blocks.append(block)

    for idx, block in enumerate(option_blocks):
        letter = letters[idx] if idx < len(letters) else letters[-1]
        update_tf_label(block, f"{letter})")


# ==================== PART 3: SHORT ANSWER ====================

def extract_short_answer_from_question(question_blocks) -> str:
    for block in question_blocks:
        txt = get_text(block)
        m = re.match(r"^\s*Đáp\s*án\s*[:\-]\s*(.+)\s*$", txt, flags=re.IGNORECASE)
        if m:
            return (m.group(1) or "").strip()
    return ""


def relabel_questions(questions):
    for i, q_blocks in enumerate(questions):
        if not q_blocks:
            continue
        update_question_label(q_blocks[0], f"Câu {i + 1}.")


def process_part(blocks, start, end, part_type):
    intro, questions = parse_questions_in_range(blocks, start, end)
    processed = []

    for q in questions:
        if part_type == "PHAN1":
            new_q, correct = shuffle_mcq_options(q)
            processed.append((new_q, correct))
        elif part_type == "PHAN2":
            new_q, key_str = shuffle_tf_options_and_key(q)
            processed.append((new_q, key_str))
        else:
            new_q = q.copy()
            ans = extract_short_answer_from_question(new_q)
            new_q = remove_short_answer_lines(new_q)  # xóa dòng Đáp án: ...
            processed.append((new_q, ans))

    shuffled_questions = shuffle_array(processed)

    relabel_questions([q for q, _ in shuffled_questions])

    if part_type == "PHAN1":
        for q, _ in shuffled_questions:
            relabel_mcq_options(q)
    elif part_type == "PHAN2":
        for q, _ in shuffled_questions:
            relabel_tf_options(q)

    result = intro.copy()
    answers = []
    for i, (q, ans) in enumerate(shuffled_questions):
        result.extend(q)
        answers.append({
            "part": 1 if part_type == "PHAN1" else (2 if part_type == "PHAN2" else 3),
            "q": i + 1,
            "answer": ans or ""
        })

    return result, answers


def shuffle_docx(file_bytes, shuffle_mode="auto", ma_de=None, ma_de_mode="full"):
    input_buffer = io.BytesIO(file_bytes)

    with zipfile.ZipFile(input_buffer, "r") as zin:
        doc_xml = zin.read("word/document.xml").decode("utf-8")
        dom = minidom.parseString(doc_xml)

        body_list = dom.getElementsByTagNameNS(W_NS, "body")
        if not body_list:
            raise Exception("Không tìm thấy w:body trong document.xml")
        body = body_list[0]

        blocks = []
        for child in body.childNodes:
            if child.nodeType == child.ELEMENT_NODE and child.localName in ["p", "tbl"]:
                blocks.append(child)

        answers_all = []

        if shuffle_mode != "auto":
            raise Exception("Bản XLSX tối ưu theo cấu trúc 3 phần. Vui lòng chọn chế độ 'Tự động (PHẦN 1,2,3)'.")

        part1_idx = find_part_index(blocks, 1)
        part2_idx = find_part_index(blocks, 2)
        part3_idx = find_part_index(blocks, 3)

        if part1_idx == -1 and part2_idx == -1 and part3_idx == -1:
            raise Exception("Không tìm thấy 'PHẦN 1/2/3'. Hãy kiểm tra lại tiêu đề phần trong Word.")

        new_blocks = []
        cursor = 0

        if part1_idx >= 0:
            new_blocks.extend(blocks[cursor:part1_idx + 1])
            cursor = part1_idx + 1
            end1 = part2_idx if part2_idx >= 0 else len(blocks)
            part1_processed, key1 = process_part(blocks, cursor, end1, "PHAN1")
            new_blocks.extend(part1_processed)
            answers_all.extend(key1)
            cursor = end1

        if part2_idx >= 0:
            new_blocks.append(blocks[part2_idx])
            start2 = part2_idx + 1
            end2 = part3_idx if part3_idx >= 0 else len(blocks)
            part2_processed, key2 = process_part(blocks, start2, end2, "PHAN2")
            new_blocks.extend(part2_processed)
            answers_all.extend(key2)
            cursor = end2

        if part3_idx >= 0:
            new_blocks.append(blocks[part3_idx])
            start3 = part3_idx + 1
            end3 = len(blocks)
            part3_processed, key3 = process_part(blocks, start3, end3, "PHAN3")
            new_blocks.extend(part3_processed)
            answers_all.extend(key3)

        other_nodes = []
        for child in list(body.childNodes):
            if child.nodeType == child.ELEMENT_NODE and child.localName not in ["p", "tbl"]:
                other_nodes.append(child)
            body.removeChild(child)

        for block in new_blocks:
            body.appendChild(block)
        for node in other_nodes:
            body.appendChild(node)

        new_xml = dom.toxml()

        output_buffer = io.BytesIO()
        with zipfile.ZipFile(output_buffer, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)

                if item.filename == "word/document.xml":
                    xml_out = new_xml
                    if ma_de is not None:
                        xml_out = replace_ma_de_placeholders(xml_out, int(ma_de), ma_de_mode)
                    zout.writestr(item, xml_out.encode("utf-8"))
                    continue

                if ma_de is not None and (
                    item.filename.startswith("word/header") or
                    item.filename.startswith("word/footer")
                ) and item.filename.endswith(".xml"):
                    try:
                        xml_in = data.decode("utf-8")
                        xml_out = replace_ma_de_placeholders(xml_in, int(ma_de), ma_de_mode)
                        data = xml_out.encode("utf-8")
                    except Exception:
                        pass

                zout.writestr(item, data)

        return output_buffer.getvalue(), answers_all


# ==================== XLSX ANSWER BUILDER ====================

def build_answer_table_xlsx(all_versions_answers, start_code=101):
    max_p = {1: 0, 2: 0, 3: 0}
    for answers in all_versions_answers:
        for r in answers:
            p = int(r.get("part", 0) or 0)
            q = int(r.get("q", 0) or 0)
            if p in max_p and q > max_p[p]:
                max_p[p] = q

    p1, p2, p3 = max_p[1], max_p[2], max_p[3]

    wb = Workbook()
    ws = wb.active
    ws.title = "Đáp án"

    header_fill = PatternFill("solid", fgColor="E6FFFA")
    group_fill = PatternFill("solid", fgColor="CCFBF1")
    thin = Side(style="thin", color="94A3B8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True)

    start_col = 2  # B
    p1_start = start_col
    p1_end = p1_start + p1 - 1
    p2_start = p1_end + 1
    p2_end = p2_start + p2 - 1
    p3_start = p2_end + 1
    p3_end = p3_start + p3 - 1

    ws["A1"] = ""
    ws["A1"].fill = group_fill
    ws["A1"].alignment = center
    ws["A1"].border = border

    if p1 > 0:
        ws.cell(row=1, column=p1_start, value="Trắc nghiệm khách quan")
        ws.merge_cells(start_row=1, start_column=p1_start, end_row=1, end_column=p1_end)
    if p2 > 0:
        ws.cell(row=1, column=p2_start, value="Trắc nghiệm đúng sai")
        ws.merge_cells(start_row=1, start_column=p2_start, end_row=1, end_column=p2_end)
    if p3 > 0:
        ws.cell(row=1, column=p3_start, value="Trắc nghiệm trả lời ngắn")
        ws.merge_cells(start_row=1, start_column=p3_start, end_row=1, end_column=p3_end)

    last_col = max(1, p3_end if p3 > 0 else (p2_end if p2 > 0 else (p1_end if p1 > 0 else 1)))
    for c in range(1, last_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = group_fill
        cell.font = bold
        cell.alignment = center
        cell.border = border

    ws["A2"] = "Mã đề"
    ws["A2"].fill = header_fill
    ws["A2"].font = bold
    ws["A2"].alignment = center
    ws["A2"].border = border

    col = start_col
    for i in range(1, p1 + 1):
        cell = ws.cell(row=2, column=col, value=f"Câu {i}")
        cell.fill = header_fill
        cell.font = bold
        cell.alignment = center
        cell.border = border
        col += 1

    for i in range(1, p2 + 1):
        cell = ws.cell(row=2, column=col, value=f"Câu {i}")
        cell.fill = header_fill
        cell.font = bold
        cell.alignment = center
        cell.border = border
        col += 1

    for i in range(1, p3 + 1):
        cell = ws.cell(row=2, column=col, value=f"Câu {i}")
        cell.fill = header_fill
        cell.font = bold
        cell.alignment = center
        cell.border = border
        col += 1

    for idx, answers in enumerate(all_versions_answers):
        code = start_code + idx
        mp = {}
        for r in answers:
            p = int(r.get("part", 0) or 0)
            q = int(r.get("q", 0) or 0)
            mp[(p, q)] = (r.get("answer", "") or "")

        rrow = 3 + idx
        ws.cell(row=rrow, column=1, value=code).alignment = center
        ws.cell(row=rrow, column=1).border = border

        col = start_col
        for q in range(1, p1 + 1):
            v = mp.get((1, q), "")
            cell = ws.cell(row=rrow, column=col, value=v)
            cell.alignment = center
            cell.border = border
            col += 1

        for q in range(1, p2 + 1):
            v = mp.get((2, q), "")
            cell = ws.cell(row=rrow, column=col, value=v)
            cell.alignment = center
            cell.border = border
            col += 1

        for q in range(1, p3 + 1):
            v = mp.get((3, q), "")
            cell = ws.cell(row=rrow, column=col, value=v)
            cell.alignment = center
            cell.border = border
            col += 1

    ws.column_dimensions["A"].width = 10
    for c in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = "B3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def create_zip_multiple(file_bytes, base_name, num_versions, shuffle_mode, ma_de_mode):
    """ZIP: nhiều đề + 1 file DAPAN_TONG_HOP.xlsx"""
    zip_buffer = io.BytesIO()
    all_versions_answers = []

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zout:
        for i in range(num_versions):
            ma_de = 101 + i
            shuffled_bytes, answers_all = shuffle_docx(
                file_bytes,
                shuffle_mode,
                ma_de=ma_de,
                ma_de_mode=ma_de_mode
            )
            all_versions_answers.append(answers_all)

            filename = f"{base_name}_V{ma_de}.docx"
            zout.writestr(filename, shuffled_bytes)

        xlsx_bytes = build_answer_table_xlsx(all_versions_answers, start_code=101)
        zout.writestr("DAPAN_TONG_HOP.xlsx", xlsx_bytes)

    return zip_buffer.getvalue()


# ==================== UI MAIN ====================

def main():
    st.markdown(
        """
<div class="hero">
  <span class="badge">Giữ nguyên MathType & OLE</span>
  <span class="badge">Tự điền mã đề trong TextBox</span>
  <span class="badge">Xuất đáp án XLSX</span>
  <h1>🎲 Trộn đề Word (3 phần) + Bảng đáp án tổng hợp (.xlsx)</h1>
  <p>Xuất nhiều mã đề 101, 102, ... • 1 file đáp án duy nhất Excel đẹp • Đề trộn xong không lộ đáp án</p>
</div>
""",
        unsafe_allow_html=True
    )

    left, right = st.columns([1.15, 0.85], gap="large")

    with left:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.subheader("Bước 1 — Chọn file Word (.docx)")
        uploaded_file = st.file_uploader(
            "Kéo thả hoặc bấm để chọn file .docx",
            type=["docx"]
        )
        if uploaded_file:
            st.success(f"✅ Đã chọn: {uploaded_file.name}")

        st.markdown('<div class="hr"></div>', unsafe_allow_html=True)

        st.subheader("Bước 2 — Thiết lập trộn")
        c1, c2 = st.columns(2)
        with c1:
            num_versions = st.number_input("Số mã đề", min_value=1, max_value=30, value=4, step=1)
        with c2:
            shuffle_mode = st.selectbox(
                "Chế độ",
                options=["auto"],
                format_func=lambda x: "Tự động (PHẦN 1,2,3)"
            )

        ma_de_mode = st.selectbox(
            "Điền {{MA_DE}} theo",
            options=["full", "2dau", "2cuoi"],
            format_func=lambda x: {
                "full": "Đầy đủ 3 số (ví dụ 101)",
                "2dau": "2 số đầu (ví dụ 10)",
                "2cuoi": "2 số cuối (ví dụ 01)"
            }[x]
        )

        st.info(
            "📌 Trong Word/TextBox:\n"
            "- Dùng **{{MA_DE}}** để điền theo lựa chọn ở trên.\n"
            "- Hoặc dùng **{{MA_DE_2DAU}}**, **{{MA_DE_2CUOI}}** nếu muốn cố định 2 số đầu/cuối."
        )

        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div style="height:10px"></div>', unsafe_allow_html=True)

        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.subheader("Bước 3 — Trộn & Tải xuống")
        run_btn = st.button("🚀 Trộn đề ngay", use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

    with right:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.subheader("Quy ước (mỗi ý 1 dòng)")
        st.markdown(
            """
- **PHẦN 1:** đáp án đúng = **gạch chân** trong nội dung phương án (sau khi trộn **sẽ bỏ gạch chân**)  
- **PHẦN 2:** gạch chân = **Đ**, không gạch chân = **S** → xuất **ĐSĐS** (sau khi trộn **sẽ bỏ gạch chân**)  
- **PHẦN 3:** đọc **`Đáp án: ...`** và **xóa dòng đáp án** khỏi đề trộn  
- File trả lời **XLSX** có **merge nhóm tiêu đề**, **kẻ bảng đẹp**
"""
        )
        st.markdown('<div class="hr"></div>', unsafe_allow_html=True)
        st.subheader("Kết quả xuất ra")
        st.markdown(
            """
- Nhiều mã đề → tải **ZIP** gồm:
  - `..._V101.docx`, `..._V102.docx`, ...
  - `DAPAN_TONG_HOP.xlsx`
"""
        )
        st.markdown('</div>', unsafe_allow_html=True)

    if run_btn:
        if not uploaded_file:
            st.error("⚠️ Thầy vui lòng chọn file .docx trước.")
            return

        try:
            with st.spinner("⏳ Đang trộn đề + điền mã đề + tạo XLSX đáp án..."):
                file_bytes = uploaded_file.read()
                base_name = uploaded_file.name.rsplit(".", 1)[0]
                base_name = re.sub(r"[^\w\s-]", "", base_name).strip() or "De"

                if num_versions == 1:
                    ma_de = 101
                    shuffled_bytes, answers_all = shuffle_docx(
                        file_bytes,
                        shuffle_mode,
                        ma_de=ma_de,
                        ma_de_mode=ma_de_mode
                    )
                    xlsx_bytes = build_answer_table_xlsx([answers_all], start_code=101)

                    st.success("✅ Hoàn tất! Đã tạo đề V101 và bảng đáp án XLSX.")

                    st.download_button(
                        label=f"📥 Tải xuống {base_name}_V101.docx",
                        data=shuffled_bytes,
                        file_name=f"{base_name}_V101.docx",
                        mime="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                        use_container_width=True
                    )
                    st.download_button(
                        label="📥 Tải xuống DAPAN_TONG_HOP.xlsx",
                        data=xlsx_bytes,
                        file_name="DAPAN_TONG_HOP.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                else:
                    zip_bytes = create_zip_multiple(
                        file_bytes,
                        base_name,
                        num_versions,
                        shuffle_mode,
                        ma_de_mode
                    )

                    st.success("✅ Hoàn tất! Đã tạo nhiều mã đề + 1 file đáp án XLSX.")

                    st.download_button(
                        label=f"📦 Tải xuống {base_name}_multi.zip",
                        data=zip_bytes,
                        file_name=f"{base_name}_multi.zip",
                        mime="application/zip",
                        use_container_width=True
                    )

        except Exception as e:
            st.error(f"❌ Lỗi: {str(e)}")

    st.markdown(
        """
<footer>
  Ngô Văn Tuấn - 0822010190
</footer>
""",
        unsafe_allow_html=True
    )


if __name__ == "__main__":
    main()
